"""
core/material_parser.py — Excel 材料清单解析 + AI列映射 + 模板生成
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List, Dict, Any, Optional
import json


# 标准模板列（6列）
TEMPLATE_COLUMNS = [
    {"key": "seq",     "label": "序号",     "aliases": ["序号", "编号", "No"]},
    {"key": "name",    "label": "工料机名称", "aliases": ["工料机名称", "材料名称", "品名", "名称", "物料名称", "设备名称"]},
    {"key": "spec",    "label": "规格型号",   "aliases": ["规格型号", "规格", "型号", "参数", "技术参数"]},
    {"key": "unit",    "label": "单位",       "aliases": ["单位", "计量单位", "单位名称"]},
    {"key": "brand",   "label": "指定品牌",   "aliases": ["指定品牌", "品牌", "推荐品牌", "参考品牌"]},
    {"key": "remark",  "label": "备注（附图）", "aliases": ["备注", "附图", "说明", "备注（附图）"]},
]


def get_column_names(file_path: str, with_row_count: bool = False):
    """获取 Excel 文件第一行的列名。

    with_row_count=True 时返回 (headers, row_count)，row_count 为数据行数（去表头）。
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    headers = [str(cell.value) if cell.value else f"(空列{chr(65+i)})" for i, cell in enumerate(ws[1])]
    row_count = ws.max_row - 1 if ws.max_row > 0 else 0
    wb.close()
    if with_row_count:
        return headers, row_count
    return headers


def auto_map_columns(headers: List[str]) -> Dict[str, Optional[int]]:
    """
    自动匹配源列 → 模板列的映射
    返回: {"name": 1, "spec": 2, "unit": 3, "brand": 4, "remark": 5, "seq": 0}
    未匹配到的为 None
    """
    mapping = {}
    for tcol in TEMPLATE_COLUMNS:
        key = tcol["key"]
        found = None
        for i, h in enumerate(headers):
            if h:
                for alias in tcol["aliases"]:
                    if alias in str(h):
                        found = i
                        break
            if found is not None:
                break
        mapping[key] = found
    return mapping


async def ai_guess_columns(headers: List[str]) -> Dict[str, Optional[int]]:
    """
    当自动匹配不够好时，用 LLM 辅助识别列映射
    """
    from extract.extractor import call_llm

    prompt = f"""你是一个Excel列映射助手。以下是用户上传的Excel文件的表头行，请将每列映射到标准模板列。

用户表头（数组，索引从0开始）：
{json.dumps(headers, ensure_ascii=False)}

标准模板列：序号、工料机名称、规格型号、单位、指定品牌、备注（附图）

请返回JSON格式的映射关系（未匹配到的填null）：
{{"seq": 索引或null, "name": 索引或null, "spec": 索引或null, "unit": 索引或null, "brand": 索引或null, "remark": 索引或null}}"""

    import json
    raw = await call_llm("primary", prompt)
    if raw:
        try:
            raw = raw.strip().lstrip("```json").rstrip("```").strip()
            return json.loads(raw)
        except json.JSONDecodeError:
            pass
    return {}


def get_preview_data(file_path: str, column_mapping: Dict[str, int], max_rows: int = 100) -> List[Dict[str, str]]:
    """
    根据列映射获取预览数据
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, max_rows + 1), values_only=True):
        data = {}
        for key, col_idx in column_mapping.items():
            if col_idx is not None and col_idx < len(row):
                val = row[col_idx]
                data[key] = str(val).strip() if val is not None else ""
            else:
                data[key] = ""
        # 跳过完全空行
        if not data.get("name", ""):
            continue
        rows.append(data)

    wb.close()
    return rows


def generate_template(output_path: str):
    """生成标准材料清单 Excel 模板（6列）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "询价表单"

    header_font = Font(name="宋体", size=9, bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_font = Font(name="宋体", size=9)
    body_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

    labels = [t["label"] for t in TEMPLATE_COLUMNS]
    for col_idx, label in enumerate(labels, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    samples = [
        ("1", "空调室外机", "制冷量206.9kW，制热量232.0kW", "台", "海尔，美的，奥克斯", ""),
        ("2", "硅酸盐水泥", "P.O42.5 袋装", "吨", "", ""),
        ("3", "钢筋 HRB400", "Φ12", "吨", "沙钢，永钢", ""),
    ]
    for row_idx, vals in enumerate(samples, 2):
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = body_font
            cell.alignment = body_align
            cell.border = thin_border

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 18

    for row_idx in range(1, len(samples) + 2):
        ws.row_dimensions[row_idx].height = 25

    wb.save(output_path)
