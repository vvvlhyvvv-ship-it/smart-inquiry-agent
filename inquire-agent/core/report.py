"""
core/report.py — V5.2 询价报告生成（Excel）

v5.0 关键变化：
- 分层输出：草稿报告（行情参考价）/ 正式报告（成交价）
- 置信度三档标注：high / medium / low
- 新增来源URL列（审计可溯源）
- AI询价说明（200字专业总结）
- 首页大字提示"AI询价建议，仅供参考"
"""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Optional
from core.router import SearchResult, SOURCE_LABEL
from utils.logger import get_logger

logger = get_logger(__name__)


# ============================================================
# 样式常量
# ============================================================

_HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# 置信度颜色
_CONF_FILL = {
    "high":   PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),  # 绿
    "medium": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),  # 黄
    "low":    PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),  # 红
}
_LOW_FONT = Font(name="微软雅黑", color="CC0000", bold=True)

# 分层标记
_DRAFT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # 草稿黄
_VERIFIED_FILL = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")  # 正式绿


# ============================================================
# 主入口：生成报告
# ============================================================

def generate_excel_report(
    results: List[SearchResult],
    output_path: str,
    project_name: str = "AI询价建议",
    report_type: str = "draft",  # draft=草稿 / formal=正式（含核实价）
    ai_summary: str = "",
) -> str:
    """
    生成 Excel 报告

    参数：
      results       : SearchResult 列表
      output_path   : 输出文件路径
      project_name  : 项目名称
      report_type   : draft=草稿报告（行情参考价）
                      formal=正式报告（含核实成交价）
      ai_summary    : AI生成的询价说明（200字）
    """
    wb = openpyxl.Workbook()

    # Sheet 1: 询价建议（主表）
    ws = wb.active
    ws.title = "询价建议"

    _write_banner(ws, report_type, project_name)
    _write_header(ws, report_type, start_row=2)
    _write_data(ws, results, report_type, start_row=3)
    _adjust_columns(ws, report_type)

    # Sheet 2: 汇总统计
    ws2 = wb.create_sheet("汇总统计")
    _write_summary(ws2, results, project_name, report_type, ai_summary)

    # Sheet 3: AI询价说明（如有）
    if ai_summary:
        ws3 = wb.create_sheet("AI询价说明")
        _write_ai_summary(ws3, ai_summary, report_type)

    wb.save(output_path)
    logger.info(f"📊 {report_type}报告已生成: {output_path}")
    return output_path


# ============================================================
# 首页大字提示横幅
# ============================================================

def _write_banner(ws, report_type: str, project_name: str):
    """第1行：大字提示横幅"""
    if report_type == "formal":
        banner = f"✅ 正式询价报告（含人工核实成交价）— {project_name}"
        fill = _VERIFIED_FILL
        font = Font(name="微软雅黑", bold=True, size=12, color="155724")
    else:
        banner = f"⚠️ AI询价建议草稿（行情参考价，未经核实，仅供参考）— {project_name}"
        fill = _DRAFT_FILL
        font = Font(name="微软雅黑", bold=True, size=12, color="856404")

    cell = ws.cell(row=1, column=1, value=banner)
    cell.font = font
    cell.fill = fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    # 合并第1行所有列
    last_col = 12 if report_type == "formal" else 11
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.row_dimensions[1].height = 28


# ============================================================
# 表头
# ============================================================

def _write_header(ws, report_type: str, start_row: int = 2):
    """写表头行"""
    if report_type == "formal":
        headers = [
            "序号", "材料名称", "规格", "单位", "供应商", "联系电话",
            "核实成交价", "价格单位", "原始单位", "换算过程", "置信度", "来源",
        ]
    else:
        headers = [
            "序号", "材料名称", "规格", "单位", "供应商", "联系电话",
            "AI建议价", "价格单位", "原始单位", "换算过程", "来源", "置信度",
        ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


# ============================================================
# 数据行
# ============================================================

def _write_data(ws, results: List[SearchResult], report_type: str, start_row: int = 3):
    """写数据行（按材料分组，合并序号/材料名/规格）"""
    row_idx = start_row
    material_seq = 0

    # 来源标签映射
    source_label = SOURCE_LABEL

    for result in results:
        if not result.success or not result.suppliers:
            # 失败行
            material_seq += 1
            ws.cell(row=row_idx, column=1, value=material_seq).border = _THIN_BORDER
            ws.cell(row=row_idx, column=2, value=result.keyword).border = _THIN_BORDER
            ws.cell(row=row_idx, column=3, value=result.spec).border = _THIN_BORDER
            ws.cell(row=row_idx, column=4, value=result.material_unit).border = _THIN_BORDER
            ws.cell(row=row_idx, column=5, value="询价失败").border = _THIN_BORDER
            err_cell = ws.cell(row=row_idx, column=6, value=result.error_message or "无结果")
            err_cell.font = Font(color="CC0000")
            err_cell.border = _THIN_BORDER
            for col in range(7, 13):
                ws.cell(row=row_idx, column=col).border = _THIN_BORDER
            row_idx += 1
            continue

        material_seq += 1
        start_merge = row_idx
        suppliers = result.suppliers

        for s in suppliers:
            sd = s.to_dict() if hasattr(s, "to_dict") else s

            # P0-12: 表头改为12列（含原始单位、换算过程）
            if report_type == "formal":
                # 核实价已下沉到供应商级：有则用核实价，无则回退原 AI 建议价
                v_price = sd.get("verified_price")
                display_price = v_price if v_price is not None else sd.get("price", 0)
                values = [
                    material_seq,
                    result.keyword,
                    result.spec,
                    result.material_unit or sd.get("unit", "").replace("元/", ""),
                    sd.get("supplier", ""),
                    sd.get("phone", ""),
                    display_price,
                    sd.get("unit", ""),
                    sd.get("unit_original", ""),
                    sd.get("convert_note", ""),
                    sd.get("confidence", "medium"),
                    source_label.get(result.source, result.source),
                ]
            else:
                values = [
                    material_seq,
                    result.keyword,
                    result.spec,
                    result.material_unit or sd.get("unit", "").replace("元/", ""),
                    sd.get("supplier", ""),
                    sd.get("phone", ""),
                    sd.get("price", 0),
                    sd.get("unit", ""),
                    sd.get("unit_original", ""),
                    sd.get("convert_note", ""),
                    source_label.get(result.source, result.source),
                    sd.get("confidence", "medium"),
                ]

            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = _THIN_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)

            # 置信度单元格上色（草稿在12列，正式在11列）
            conf_col = 12 if report_type == "formal" else 11
            conf = sd.get("confidence", "medium")
            conf_cell = ws.cell(row=row_idx, column=conf_col)
            conf_cell.fill = _CONF_FILL.get(conf, _CONF_FILL["medium"])
            if conf == "low":
                conf_cell.font = _LOW_FONT

            row_idx += 1

        # 合并序号、材料名称、规格、单位
        end_merge = row_idx - 1
        if end_merge > start_merge:
            for col in [1, 2, 3, 4]:
                ws.merge_cells(
                    start_row=start_merge, start_column=col,
                    end_row=end_merge, end_column=col,
                )
                for r in range(start_merge, end_merge + 1):
                    ws.cell(row=r, column=col).alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )


# ============================================================
# 列宽
# ============================================================

def _adjust_columns(ws, report_type: str):
    """调整列宽"""
    if report_type == "formal":
        widths = [6, 26, 24, 7, 30, 15, 10, 12, 10, 30, 8, 14]
    else:
        widths = [6, 26, 24, 7, 30, 15, 10, 12, 10, 30, 14, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# 汇总统计
# ============================================================

def _write_summary(ws, results, project_name, report_type, ai_summary):
    """汇总统计 Sheet"""
    total = len(results)
    success_count = sum(1 for r in results if r.success)

    # 置信度统计
    conf_counts = {"high": 0, "medium": 0, "low": 0}
    source_counts = {}
    for r in results:
        source_counts[r.source] = source_counts.get(r.source, 0) + 1
        for s in (r.suppliers or []):
            sd = s.to_dict() if hasattr(s, "to_dict") else s
            conf = sd.get("confidence", "medium")
            conf_counts[conf] = conf_counts.get(conf, 0) + 1

    source_label = SOURCE_LABEL
    source_text = ", ".join(
        f"{source_label.get(k, k)} {v}项" for k, v in source_counts.items() if v
    )

    rows = [
        ("项目名称", project_name),
        ("报告类型", "正式报告（含核实成交价）" if report_type == "formal" else "AI建议草稿（行情参考价）"),
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("材料总数", f"{total} 项"),
        ("询价成功", f"{success_count} 项（{success_count*100//total if total else 0}%）"),
        ("置信度分布", f"高 {conf_counts['high']} / 中 {conf_counts['medium']} / 低 {conf_counts['low']}"),
        ("价格来源", source_text),
    ]
    if report_type == "formal":
        verified_count = sum(1 for r in results if getattr(r, "verified", False))
        rows.append(("已核实", f"{verified_count} 项"))

    for i, (k, v) in enumerate(rows, 1):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)


# ============================================================
# AI 询价说明
# ============================================================

def _write_ai_summary(ws, ai_summary: str, report_type: str):
    """AI 询价说明 Sheet"""
    title = "AI询价说明" + ("（正式报告）" if report_type == "formal" else "（草稿）")
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ws.cell(row=4, column=1, value=ai_summary).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 80
    ws.row_dimensions[4].height = 120

    # 免责声明
    disclaimer_row = 6
    if report_type == "draft":
        ws.cell(row=disclaimer_row, column=1, value=(
            "⚠️ 免责声明：本报告价格均为AI生成的行情参考价，未经人工核实，"
            "不可直接作为最终采购或审计依据。请造价人员电话核实后使用。"
        )).font = Font(color="CC0000", bold=True)
    else:
        ws.cell(row=disclaimer_row, column=1, value=(
            "✅ 本报告含人工核实成交价，可作为正式询价依据。"
            "未核实项仍标注为行情参考价，请注意区分。"
        )).font = Font(color="155724", bold=True)
